# -*- coding: utf-8 -*-
"""
Created on Sat Oct 28 23:09:57 2017

@author: goldw
"""

from openpyxl import load_workbook
from random import randint

wb = load_workbook("C:\\Users\\goldw\Documents\\\MajorPathData8.xlsx")
wb2 = load_workbook("C:\\Users\\goldw\Documents\\\majorslist.xlsx")
ws = wb.active

#fill zeros

#INTERESTS
interest_dict = {
        "Commercial Art & Design1" : ['Social Science', 'Advertising', 'Art'],
        "Commercial Art & Design2" : ['Consumer Behavior', 'Art', 'Technology'],
        "Commercial Art & Design3" : ['Entrepreneurship', 'Art', 'Graphic Design'],
        "Creative Advertising1" : ['Consumer Beahvoir', 'Creative Writing', 'Advertising'],
        "Creative Advertising2" : ['Marketing', 'Business Development', 'Entrepreneurship'],
        "Creative Advertising3" : ['Art', 'Market Research', 'Management'],
        "Accounting1" : ['Critical Thinking','Business Development','Mathematics'],
        "Accounting2" : ['Business Development','Technology','Mathematics'],
        "Accounting3" : ['Business Development','Data Analytics','Public Service'],
        "Business Management & Administration1" : ['Critical Thinking','Management','Business Development'],
        "Business Management & Administration2" : ['Marketing','Market Research','Entrepreneurship'],
        "Business Management & Administration3" : ['Technology','Marketing','Business Development'],
        "Economics1" : ['Mathematics','Critical Thinking','Data Analytics'],
        "Economics2" : ['Mathematics','Social Science','Market Research'],
        "Economics3" : ['Business Development','Consumer Behavior','Mathematics'],
        "Marketing1" : ['Market Research','Advertising','Marketing'],
        "Marketing2" : ['Creative Writing','Graphic Design','Data Analytics'],
        "Marketing3" : ['Entrepreneurship','Marketing','Consumer Behavoir',],
        "Finance1" : ['Mathematics','Problem Solving','Management'],
        "Finance2" : ['Critical Thinking','Mathematics','Business Development'],
        "Finance3" : ['Mathematics','Entrepreneurship','Computer Science'],
        "International Business1" : ['Management','Problem Solving','Critical Thinking'],
        "International Business2" : ['Data Analytics','Consumer Behavior','Market Research'],
        "International Business3" : ['Political Science','Entrepreneurship','Business Development'],
        "Business1" : ['Consumer Behavior','Managment','Market Research'],
        "Business2" : ['Business Development','Critical Thinking','Problem Solving'],
        "Business3" : ['Business Development','Marketing','Management'],
        "Visual & Performing Arts1" : ['Live Performance','Art','Film & Media'],
        "Visual & Performing Arts2" : ['Film & Media','Graphic Design','Music'],
        "Visual & Performing Arts3" : ['Creative Writing','Art','Live Performance'],
        "Music Production1" : ['Music','Art','Live Performance'],
        "Music Production2" : ['Music','Public Service','Live Performance'],
        "Music Production3" : ['Music','Art','Film & Media'],
        "Communication1" : ['Data Analytics','Market Research','Consumer Behavior'],
        "Communication2" : ['Social Science','Creative Writing','Consumer Behavior'],
        "Communication3" : ['Creative Writing','Advertising','Marketing'],
        "Journalism1" : ['Creative Writing','Social Science','Advertising'],
        "Journalism2" : ['Creative Writing','Entrepreneurship','Literature'],
        "Journalism3" : ['Literature','Public Service','History'],
        "Public Relations1" : ['Advertising','Business Development','Consumer Behavior'],
        "Public Relations2" : ['Creative Writing','History','Marketing'],
        "Public Relations3" : ['Political Science','Creative Writing','Critical Thinking'],
        "Entrepreneurship1" : ['Business Development','Management','Entrepreneurship'],
        "Entrepreneurship2" : ['Advertising','Market Research','Entrepreneurship'],
        "Entrepreneurship3" : ['Entrepreneurship','Technology','Marketing'],
        "Legal Studies1" : ['Social Science','Political Science','Problem Solving'],
        "Legal Studies2" : ['Public Service','Critical Thinking','Entrepreneurship'],
        "Legal Studies3" : ['Literature','History','Data Analytics'],
        "Real Estate1" : ['Managment','Architecture','Market Research'],
        "Real Estate2" : ['Public Service','Data Analytics','Technology'],
        "Real Estate3" : ['Marketing','Business Development','Architecture'],
        "Sports Administration1" : ['Management','Exercise Physiology','Medicine'],
        "Sports Administration2" : ['Critical Thinking','Problem Solving','Market Research'],
        "Sports Administration3" : ['Public Service','Consumer Behavior','Social Science'],
        "Architecture1" : ['Architecture','Management','Data Analytics'],
        "Architecture2" : ['Critical Thinking','Enviromental Awareness','Public Service'],
        "Architecture3" : ['Art','Architecture','Technology'],
        "Sports Therapy1" : ['Exercise Physiology','Technology','Data Analytics'],
        "Sports Therapy2" : ['Entrepreneurship','Public Service','Business Development'],
        "Sports Therapy3" : ['Science','Management','Advertising'],
        "Political Science1" : ['Political Science','Social Science','History'],
        "Political Science2" : ['Public Service','Political Science','Consumer Behavior'],
        "Political Science3" : ['Data Analytics','Problem Solving','Political Science'],
        "Human Resources1" : ['Social Science','Management','Business Development'],
        "Human Resources2" : ['Social Science','Public Service','Management'],
        "Human Resources3" : ['Problem Solving','Consumer Behavior','Public Service'],
        "Communication Technology1" : ['Market Research','Data Analytics','Technology'],
        "Communication Technology2" : ['Advertising','Management','Computer Science'],
        "Communication Technology3" : ['Business Development','Consumer Behavior','Graphic Design'],
        "Data Analytics1" : ['Data Analytics','Mathematics','Market Research'],
        "Data Analytics2" : ['Computer Science','Data Analytics','Consumer Behavior'],
        "Data Analytics3" : ['Data Analytics','Critical Thinking','Problem Solving'],
        "Public Health1" : ['Public Service','Medicine','Critical Thinking'],
        "Public Health2" : ['Social Science','Medicine','Political Science'],
        "Public Health3" : ['Medicine','Political Science','Public Service'],
        "Kinesology1" : ['Exercise Physiology','Public Service','Medicine'],
        "Kinesology2" : ['Science','Medicine','Exercise Physiology'],
        "Kineesology3" : ['Exercise Physiology','Critical Thinking','Medicine'],
        "English Literature1" : ['Creative Writing','Literature','History'],
        "English Literature2" : ['Literature','Live Performance','Creative Writing'],
        "English Literature3" : ['Literature','History','Critical Thinking'],
        "International Studies1" : ['Hisotry','Political Science','Literature'],
        "International Studies2" : ['Hisotry','Public Service','Political Science'],
        "International Studies3" : ['Political Science','Business Development','History'],
        "History1" : ['Creative Writing','Political Science','History'],
        "History2" : ['Public Service','Data Analytics','Art'],
        "History3" : ['History','Literature','Political Science'],
        "Public Policy1" : ['Business Development','Public Service','Social Science'],
        "Public Policy2" : ['Management','Literature','Enviromental Awareness'],
        "Public Policy3" : ['Political Science','History','Consumer Behavior'],
        "Operations Logistics1" : ['Mathematics','Data Analytics','Critical Thinking'],
        "Operations Logistics2" : ['Problem Solving','Business Development','Engineering'],
        "Operations Logistics3" : ['Market Research','Marketing','Critical Thinking'],
        "Information Systems1" : ['Technology','Mathematics','Computer Science'],
        "Information Systems2" : ['Data Analytics','Problem Solving','Critical Thinking'],
        "Information Systems3" : ['Consumer Behavior','Engineering','Business Development'],
        "Educational Administration1" : ['Public Service','Literature','History'],
        "Educational Administration2" : ['Social Science','Management','Creativing Writing'],
        "Educational Administration3" : ['Management','Public Service','Social Science'],
        "Human Services1" : ['Public Service','Political Science','Business Development'],
        "Human Services2" : ['Enviromental Awareness','Management','Public Service'],
        "Human Services3" : ['Literature','Data Analytics','Public Service'],
        "Fine Arts1" : ['Live Performance','Music','Art'],
        "Fine Arts2" : ['Art','Film & Media','History'],
        "Fine Arts3" : ['Creative Writing','Graphic Design','Live Performance'],
        "Health Services1" : ['Medicine','Consumer Behavior','Science'],
        "Health Services2" : ['Medicine','Public Service','Business Development'],
        "Health Services3" : ['Medicine','Market Research','Public Service'],
        "Nursing1" : ['Public Service','Science','Medicine'],
        "Nursing2" : ['Exercise Physiology','Social Science','Public Service'],
        "Nursing3" : ['Medicine','Public Service','Management'],
        "Philosophy1" : ['Creative Writing','Critical Thinking','Problem Solving'],
        "Philosophy2" : ['Social Science','Political Science','Literature'],
        "Philosophy3" : ['Art','History','Critical Thinking'],
        "Art History1" : ['History','Architecture','Literature'],
        "Art History2" : ['Art','History','Music'],
        "Art History3" : ['Creative Writing','Live Performance','Film & Media'],
        "Sociology1" : ['Social Science','Public Service','Data Analytics'],
        "Sociology2" : ['History','Political Science','Literature'],
        "Sociology3" : ['Social Science','Literature','Data Analytics'],
        "U.S. History1" : ['History','Political Science','Social Science'],
        "U.S. History2" : ['History','Literature','Consumer Behavior'],
        "U.S. History3" : ['History','Literature','Political Science'],
        "Liberal Arts1" : ['Art','Literature','Creative Writing'],
        "Liberal Arts2" : ['Public Service','Creative Writing','Social Science'],
        "Liberal Arts3" : ['Literature','Creative Writing','Advertising'],
        "Linguistics1" : ['Literature','History','Creative Writing'],
        "Linguistics2" : ['Literature','Art','History'],
        "Linguistics3" : ['Literature','Social Service','Science'],
        "Film1" : ['Film & Media','Music','Art'],
        "Film2" : ['Film & Media','Public Service','Entrepeneurship'],
        "Film3" : ['Film & Media','Art','Graphic Design'],
        "Theatre1" : ['Live Performance','Music','Art'],
        "Theatre2" : ['Public Service','Live Performance','Film & Media'],
        "Theatre3" : ['Art','Creative Writing','Literature'],
        "Studio Art1" : ['Art','Film & Media','Graphic Design'],
        "Studio Art2" : ['Live Performance','Technology','Art'],
        "Studio Art3" : ['Art','Graphic Design','Technology'],
        "Student Counseling1" : ['Public Service','Social Science','Critical Thinking'],
        "Student Counseling2" : ['Public Service','Social Science','Management'],
        "Student Counseling3" : ['Social Science','Literature','Problem Solving'],
        "Computer Administration and Security1" : ['Technology','Business Development','Management'],
        "Computer Administration and Security2" : ['Data Analytics','Computer Science','Technology'],
        "Computer Administration and Security3" : ['Business Development','Mathematics','Technology'],
        "Mathematics1" : ['Technology','Mathematics','Computer Science'],
        "Mathematics2" : ['Critical Thinking','Data Analytics','Problem Solving'],
        "Mathematics3" : ['Engineering','Business Development','Mathematics'], #163
        "Physics1" : ['Critical Thinking','Science','Mathematics'],
        "Physics2" : ['Data Analytics','Computer Science','Science'],
        "Physics3" : ['Technology','Engineering','Critical Thinking'],
        "Marine Science1" : ['Science','Environmental Awareness','Mathematics'],
        "Marine Science2" : ['Environmental Awareness','Public Service','Science'],
        "Marine Science3" : ['Environmental Awareness','Technology','Public Service'],
        "Petroleum Engineering1" : ['Engineering','Environmental Awareness','Science'],
        "Petroleum Engineering2" : ['Engineering','Business Development','Critical Thinking'],
        "Petroleum Engineering3" : ['Engineering','Hisotry','Science'],
        "Materials Engineering1" : ['Engineering','Science','Consumer Behavior'],
        "Materials Engineering2" : ['Engineering','Market Research','Technology'],
        "Materials Engineering3" : ['Engineering','Art','Science'],
        "Industrial Engineering1" : ['Engineering','Consumer Behavior','Market Research'],
        "Industrial Engineering2" : ['Engineering','Data Analytics','Mathematics'],
        "Industrial Engineering3" : ['Engineering','Business Development','Marketing'],
        "Environmental Engineering1" : ['Engineering','Environmental Awareness','Technology'],
        "Environmental Engineering2" : ['Engineering','Public Service','Environmental Awareness'],
        "Environmental Engineering3" : ['Engineering','Environmental Awareness','Science'],
        "Electrical Engineering1" : ['Engineering','Science','Computer Science'],
        "Electrical Engineering2" : ['Engineering','Technology','Mathematics'],
        "Electrical Engineering3" : ['Engineering','Mathematics','Computer Science'],
        "Civil Engineering1" : ['Engineering','Architecture','Mathematics'],
        "Civil Engineering2" : ['Engineering','Public Service','Problem Solving'],
        "Civil Engineering3" : ['Engineering','Architecture','Environmental Awareness'],
        "Biomedical Engineering1" : ['Engineering','Medicine','Public Service'],
        "Biomedical Engineering2" : ['Engineering','Medicine','Science'],
        "Biomedical Engineering3" : ['Engineering','Medicine','Problem Solving'],
        "Aerospace Engineering1" : ['Engineering','Science','Technology'],
        "Aerospace Engineering2" : ['Engineering','Critical Thinking','Technology'],
        "Aerospace Engineering3" : ['Engineering','Mathematics','Science'],
        "" : ['','',''],
        "" : ['','',''],
        "" : ['','',''],
        }
     
#extrapolation, major = 2-176
#major #2-99
def majorfunc(major, interest1, interest2, interest3):
    if interest1 == "Creative Writing":
        for row in ws.iter_cols(min_row =major, max_row = major, min_col = 32, max_col = 40):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major, max_row = major, min_col = 2, max_col = 2):
            for cell in row:
                cell.value = 1
    elif interest1 == "Critical Thinking":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 41, max_col = 49):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 3, max_col = 3):
            for cell in row:
                cell.value = 1
    elif interest1 == "Problem Solving":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 50, max_col = 58):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 4, max_col = 4):
            for cell in row:
                cell.value = 1
    elif interest1 == "Public Service":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 59, max_col = 67):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 5, max_col = 5):
            for cell in row:
                cell.value = 1
    elif interest1 == "Live Performance":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 68, max_col = 76):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 6, max_col = 6):
            for cell in row:
                cell.value = 1
    elif interest1 == "Technology":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 77, max_col = 85):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 7, max_col = 7):
            for cell in row:
                cell.value = 1
    elif interest1 == "Data Analytics":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 86, max_col = 94):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 8, max_col = 8):
            for cell in row:
                cell.value = 1
    elif interest1 == "Consumer Behavior":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 95, max_col = 103):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 9, max_col = 9):
            for cell in row:
                cell.value = 1
    elif interest1 == "Mathematics":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 104, max_col = 112):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 10, max_col = 10):
            for cell in row:
                cell.value = 1
    elif interest1 == "Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 113, max_col = 121):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 11, max_col = 11):
            for cell in row:
                cell.value = 1
    elif interest1 == "Engineering":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 122, max_col = 130):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 12, max_col = 12):
            for cell in row:
                cell.value = 1
    elif interest1 == "Business Development":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 131, max_col = 139):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 13, max_col = 13):
            for cell in row:
                cell.value = 1
    elif interest1 == "Entrepeneurship":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 140, max_col = 148):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 14, max_col = 14):
            for cell in row:
                cell.value = 1
    elif interest1 == "Environmental Awareness":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 149, max_col = 157):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 15, max_col = 15):
            for cell in row:
                cell.value = 1
    elif interest1 == "Management":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 158, max_col = 166):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 16, max_col = 16):
            for cell in row:
                cell.value = 1
    elif interest1 == "Advertising":
        for row in ws.iter_cols(min_row =major, max_row = major, min_col = 167, max_col = 175):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 17, max_col = 17):
            for cell in row:
                cell.value = 1
    elif interest1 == "Marketing":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 176, max_col = 184):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 18, max_col = 18):
            for cell in row:
                cell.value = 1
    elif interest1 == "Music":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 185, max_col = 193):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 19, max_col = 19):
            for cell in row:
                cell.value = 1
    elif interest1 == "Art":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 194, max_col = 202):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 20, max_col = 20):
            for cell in row:
                cell.value = 1
    elif interest1 == "Market Research":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 203, max_col = 211):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 21, max_col = 21):
            for cell in row:
                cell.value = 1
    elif interest1 == "Medicine":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 212, max_col = 220):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 22, max_col = 22):
            for cell in row:
                cell.value = 1
    elif interest1 == "Film & Media":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 221, max_col = 229):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 23, max_col = 23):
            for cell in row:
                cell.value = 1
    elif interest1 == "Graphic Design":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 230, max_col = 238):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 24, max_col = 24):
            for cell in row:
                cell.value = 1
    elif interest1 == "Computer Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 239, max_col = 247):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 25, max_col = 25):
            for cell in row:
                cell.value = 1
    elif interest1 == "Exercise Physiology":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 248, max_col = 256):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 26, max_col = 26):
            for cell in row:
                cell.value = 1
    elif interest1 == "History":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 257, max_col = 265):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 27, max_col = 27):
            for cell in row:
                cell.value = 1
    elif interest1 == "Architecture":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 266, max_col = 274):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 28, max_col = 28):
            for cell in row:
                cell.value = 1
    elif interest1 == "Literature":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 275, max_col = 283):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 29, max_col = 29):
            for cell in row:
                cell.value = 1
    elif interest1 == "Political Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 284, max_col = 292):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 30, max_col = 30):
            for cell in row:
                cell.value = 1
    elif interest1 == "Social Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 293, max_col = 301):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 31, max_col = 31):
            for cell in row:
                cell.value = 1
def majorfunc2(major, interest1, interest2, interest3):
    if interest2 == "Creative Writing":
        for row in ws.iter_cols(min_row =major, max_row = major, min_col = 32, max_col = 40):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major, max_row = major, min_col = 2, max_col = 2):
            for cell in row:
                cell.value = 1
    elif interest2 == "Critical Thinking":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 41, max_col = 49):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 3, max_col = 3):
            for cell in row:
                cell.value = 1
    elif interest2 == "Problem Solving":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 50, max_col = 58):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 4, max_col = 4):
            for cell in row:
                cell.value = 1
    elif interest2 == "Public Service":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 59, max_col = 67):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 5, max_col = 5):
            for cell in row:
                cell.value = 1
    elif interest2 == "Live Performance":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 68, max_col = 76):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 6, max_col = 6):
            for cell in row:
                cell.value = 1
    elif interest2 == "Technology":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 77, max_col = 85):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 7, max_col = 7):
            for cell in row:
                cell.value = 1
    elif interest2 == "Data Analytics":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 86, max_col = 94):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 8, max_col = 8):
            for cell in row:
                cell.value = 1
    elif interest2 == "Consumer Behavior":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 95, max_col = 103):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 9, max_col = 9):
            for cell in row:
                cell.value = 1
    elif interest2 == "Mathematics":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 104, max_col = 112):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 10, max_col = 10):
            for cell in row:
                cell.value = 1
    elif interest2 == "Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 113, max_col = 121):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 11, max_col = 11):
            for cell in row:
                cell.value = 1
    elif interest2 == "Engineering":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 122, max_col = 130):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 12, max_col = 12):
            for cell in row:
                cell.value = 1
    elif interest2 == "Business Development":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 131, max_col = 139):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 13, max_col = 13):
            for cell in row:
                cell.value = 1
    elif interest2 == "Entrepeneurship":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 140, max_col = 148):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 14, max_col = 14):
            for cell in row:
                cell.value = 1
    elif interest2 == "Environmental Awareness":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 149, max_col = 157):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 15, max_col = 15):
            for cell in row:
                cell.value = 1
    elif interest2 == "Management":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 158, max_col = 166):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 16, max_col = 16):
            for cell in row:
                cell.value = 1
    elif interest2 == "Advertising":
        for row in ws.iter_cols(min_row =major, max_row = major, min_col = 167, max_col = 175):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 17, max_col = 17):
            for cell in row:
                cell.value = 1
    elif interest2 == "Marketing":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 176, max_col = 184):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 18, max_col = 18):
            for cell in row:
                cell.value = 1
    elif interest2 == "Music":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 185, max_col = 193):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 19, max_col = 19):
            for cell in row:
                cell.value = 1
    elif interest2 == "Art":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 194, max_col = 202):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 20, max_col = 20):
            for cell in row:
                cell.value = 1
    elif interest2 == "Market Research":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 203, max_col = 211):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 21, max_col = 21):
            for cell in row:
                cell.value = 1
    elif interest2 == "Medicine":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 212, max_col = 220):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 22, max_col = 22):
            for cell in row:
                cell.value = 1
    elif interest2 == "Film & Media":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 221, max_col = 229):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 23, max_col = 23):
            for cell in row:
                cell.value = 1
    elif interest2 == "Graphic Design":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 230, max_col = 238):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 24, max_col = 24):
            for cell in row:
                cell.value = 1
    elif interest2 == "Computer Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 239, max_col = 247):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 25, max_col = 25):
            for cell in row:
                cell.value = 1
    elif interest2 == "Exercise Physiology":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 248, max_col = 256):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 26, max_col = 26):
            for cell in row:
                cell.value = 1
    elif interest2 == "History":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 257, max_col = 265):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 27, max_col = 27):
            for cell in row:
                cell.value = 1
    elif interest2 == "Architecture":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 266, max_col = 274):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 28, max_col = 28):
            for cell in row:
                cell.value = 1
    elif interest2 == "Literature":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 275, max_col = 283):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 29, max_col = 29):
            for cell in row:
                cell.value = 1
    elif interest2 == "Political Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 284, max_col = 292):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 30, max_col = 30):
            for cell in row:
                cell.value = 1
    elif interest2 == "Social Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 293, max_col = 301):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 31, max_col = 31):
            for cell in row:
                cell.value = 1
def majorfunc3(major, interest1, interest2, interest3):
    if interest3 == "Creative Writing":
        for row in ws.iter_cols(min_row =major, max_row = major, min_col = 32, max_col = 40):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major, max_row = major, min_col = 2, max_col = 2):
            for cell in row:
                cell.value = 1
    elif interest3 == "Critical Thinking":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 41, max_col = 49):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 3, max_col = 3):
            for cell in row:
                cell.value = 1
    elif interest3 == "Problem Solving":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 50, max_col = 58):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 4, max_col = 4):
            for cell in row:
                cell.value = 1
    elif interest3 == "Public Service":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 59, max_col = 67):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 5, max_col = 5):
            for cell in row:
                cell.value = 1
    elif interest3 == "Live Performance":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 68, max_col = 76):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 6, max_col = 6):
            for cell in row:
                cell.value = 1
    elif interest3 == "Technology":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 77, max_col = 85):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 7, max_col = 7):
            for cell in row:
                cell.value = 1
    elif interest3 == "Data Analytics":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 86, max_col = 94):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 8, max_col = 8):
            for cell in row:
                cell.value = 1
    elif interest3 == "Consumer Behavior":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 95, max_col = 103):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 9, max_col = 9):
            for cell in row:
                cell.value = 1
    elif interest3 == "Mathematics":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 104, max_col = 112):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 10, max_col = 10):
            for cell in row:
                cell.value = 1
    elif interest3 == "Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 113, max_col = 121):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 11, max_col = 11):
            for cell in row:
                cell.value = 1
    elif interest3 == "Engineering":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 122, max_col = 130):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 12, max_col = 12):
            for cell in row:
                cell.value = 1
    elif interest3 == "Business Development":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 131, max_col = 139):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 13, max_col = 13):
            for cell in row:
                cell.value = 1
    elif interest3 == "Entrepeneurship":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 140, max_col = 148):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 14, max_col = 14):
            for cell in row:
                cell.value = 1
    elif interest3 == "Environmental Awareness":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 149, max_col = 157):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 15, max_col = 15):
            for cell in row:
                cell.value = 1
    elif interest3 == "Management":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 158, max_col = 166):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 16, max_col = 16):
            for cell in row:
                cell.value = 1
    elif interest3 == "Advertising":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 167, max_col = 175):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 17, max_col = 17):
            for cell in row:
                cell.value = 1
    elif interest3 == "Marketing":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 176, max_col = 184):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 18, max_col = 18):
            for cell in row:
                cell.value = 1
    elif interest3 == "Music":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 185, max_col = 193):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 19, max_col = 19):
            for cell in row:
                cell.value = 1
    elif interest3 == "Art":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 194, max_col = 202):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 20, max_col = 20):
            for cell in row:
                cell.value = 1
    elif interest3 == "Market Research":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 203, max_col = 211):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 21, max_col = 21):
            for cell in row:
                cell.value = 1
    elif interest3 == "Medicine":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 212, max_col = 220):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 22, max_col = 22):
            for cell in row:
                cell.value = 1
    elif interest3 == "Film & Media":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 221, max_col = 229):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 23, max_col = 23):
            for cell in row:
                cell.value = 1
    elif interest3 == "Graphic Design":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 230, max_col = 238):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 24, max_col = 24):
            for cell in row:
                cell.value = 1
    elif interest3 == "Computer Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 239, max_col = 247):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 25, max_col = 25):
            for cell in row:
                cell.value = 1
    elif interest3 == "Exercise Physiology":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 248, max_col = 256):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 26, max_col = 26):
            for cell in row:
                cell.value = 1
    elif interest3 == "History":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 257, max_col = 265):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 27, max_col = 27):
            for cell in row:
                cell.value = 1
    elif interest3 == "Architecture":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 266, max_col = 274):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 28, max_col = 28):
            for cell in row:
                cell.value = 1
    elif interest3 == "Literature":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 275, max_col = 283):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 29, max_col = 29):
            for cell in row:
                cell.value = 1
    elif interest3 == "Political Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 284, max_col = 292):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 30, max_col = 30):
            for cell in row:
                cell.value = 1
    elif interest3 == "Social Science":
        for row in ws.iter_cols(min_row =major,max_row = major, min_col = 293, max_col = 301):
            for cell in row:
                cell.value = randint(0,1)
        for row in ws.iter_cols(min_row = major,max_row = major, min_col = 31, max_col = 31):
            for cell in row:
                cell.value = 1
                

def interp():
    z = 0
    b = 0
    doe = list(interest_dict.keys()) #majors
    doe2 = list(interest_dict.values())
     #interests
    min_rowz = 1
    while(z<60): # number of majors + 1, 64
        max_rowz = min_rowz + 332
        for col in ws.iter_cols(min_row= min_rowz, max_col= 1, max_row = max_rowz): #writes major names to excel file, max_row = size of dataset
            for cell in col:
                cell.value = doe[z]
                wb.save("C:\\Users\\goldw\\Documents\\MajorPathData7.xlsx")
                min_rowz = min_rowz + 1
        z = z + 1
        print(z)
    min_rowz = 1
    while(b < 60): # number of majors
        x = 0
        max_rowz = min_rowz + 332
        while(x <333): #each major subsection gets 333 entries in the dataset
            majorfunc(major = min_rowz,interest1 = doe2[b][0], interest2 = doe2[b][1], interest3 = doe2[b][2])
            majorfunc2(major = min_rowz,interest1 = doe2[b][0], interest2 = doe2[b][1], interest3 = doe2[b][2])
            majorfunc3(major = min_rowz,interest1 = doe2[b][0], interest2 = doe2[b][1], interest3 = doe2[b][2])
            wb.save("C:\\Users\\goldw\\Documents\\MajorPathData7.xlsx")
            x = x + 1
            min_rowz = min_rowz + 1
        b = b + 1
        print(b)

def fill_zeros():
    a = 0
    for col in ws.iter_cols(min_row = 1, max_row = 19980, min_col = 2, max_col = 2):
        for cell in col:
            print(a)
            cell.internal_value == 'None'
            cell.value = 0
            wb.save("C:\\Users\\goldw\\Documents\\MajorPathData8.xlsx")
            a = a+1
            
            
fill_zeros()
                



    